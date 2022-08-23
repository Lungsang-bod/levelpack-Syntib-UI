from lark import Tree
from openpyxl import Workbook


def larktree2treespread(trees, out_file, mode='xlsx'):
    # a. prepare data
    ts_sheets = []
    for i, tree in enumerate(trees):
        # 1. get tree and leaves
        tree_struct, leaves = [], []

        transform_tree(tree, 1, '\t', tree_struct, leaves)
        square_bracket_struct(tree_struct)
        lower_terminal_nodes(tree_struct)

        # 2. create sheet
        # add initial column
        for n, line in enumerate(tree_struct):
            tree_struct[n] = [''] + line
        tree_struct[-1][0] = 'P'
        leaves = ['W'] + leaves

        # constitute sheet
        sheet = []
        sheet.extend(tree_struct)
        sheet.extend([leaves])
        sheet.extend(10 * [[''] + leaves[1:]])

        ts_sheets.append((i, sheet))

    # b. write to outfile
    if mode == 'xlsx':
        workbook = Workbook()
        workbook.remove_sheet(workbook.active)
        for name, sheet_data in ts_sheets:
            sheet = workbook.create_sheet(title=str(name))
            for r, row in enumerate(sheet_data):
                for c, cell in enumerate(row):
                    sheet.cell(row=r+1, column=c+1, value=cell)
        workbook.save(out_file)
    else:
        raise NotImplementedError


def transform_tree(tree, level, indent_str, prettied, leaves):
    if len(tree.children) == 1 and not isinstance(tree.children[0], Tree):
        # add missing level
        longest_level = max([len(p) for p in prettied])
        if len(prettied) < level:
            to_add = [''] * (longest_level - 1) if longest_level >= 2 else []
            prettied.append(to_add)
        if len(prettied) < level + 1:
            to_add = [''] * (longest_level - 1) if longest_level >= 2 else []
            prettied.append(to_add)

        # add nodes + leaves
        try:
            prettied[level-1].append((tree.data.value, len(tree.children)))
        except AttributeError:
            prettied[level-1].append((tree.data, len(tree.children)))
        prettied[level].append(tree.children[0].type + '*')
        leaves.append(tree.children[0].value)

        # level up structure horizontally
        level_structure_horizontally(level, prettied)

        return

    # add missing level
    if len(prettied) < level:
        longest_level = max([len(p) for p in prettied]) if prettied else 0
        to_add = [''] * (longest_level - 1) if longest_level >= 2 else []
        prettied.append(to_add)

    # add node
    try:
        prettied[level-1].append((tree.data.value, len(tree.children)))
    except AttributeError:
        prettied[level - 1].append((tree.data, len(tree.children)))

    for n in tree.children:
        if isinstance(n, Tree):
            transform_tree(n, level + 1, indent_str, prettied, leaves)
        else:
            while len(prettied) - 1 < level:
                longest_level = max([len(p) for p in prettied])
                to_add = [''] * (longest_level - 1) if longest_level >= 2 else []
                prettied.append(to_add)

            # add node and leave
            prettied[level].append(n.type + '*')
            leaves.append(n.value)

            # level up structure horizontally
            level_structure_horizontally(level, prettied)


def level_structure_horizontally(level, structure):
    size = len(structure[level])
    for p in structure:
        while len(p) < size:
            p.append('')


def square_bracket_struct(struct):
    for i, level in enumerate(struct):
        for j, cell in enumerate(level):
            if isinstance(cell, tuple):
                content, size = cell
                end = 0
                while size > 0:
                    for k, el in enumerate(struct[i+1]):
                        if k < j:
                            continue

                        end += 1
                        if el != '':
                            size -= 1
                        if size <= 0:
                            break
                struct[i][j] = f'[{cell[0]}'
                struct[i][j + end-1] += ']'


def lower_terminal_nodes(struct):
    for i, level in enumerate(struct):
        for j, cell in enumerate(level):
            if '*' in cell:
                struct[len(struct)-1][j] = f'[{cell[:-1]}]'
                if i < len(struct) - 1:
                    struct[i][j] = ''
